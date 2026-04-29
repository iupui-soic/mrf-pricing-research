#!/usr/bin/env python3
"""
extract_npi_from_mrfs.py
========================
Pulls hospital NPI from each downloaded MRF file (we already have 481 of
them) so the crosswalk doesn't need a CMS NPPES download for v2.

CMS v2.0 / v3.0 MRF spec puts hospital NPI in `type_2_npi`:
  - JSON: top-level `type_2_npi` (array of NPI strings).
  - CSV: row 1 has metadata column headers including `type_2_npi`; row 2
    has corresponding values. Pipe-delimited if multi-location.

Older v1.x files sometimes use `hospital_npi` or just `npi`; we try those
as fallbacks.

Output: /data0/mrf-pricing-research/crosswalk/ccn_to_npi.csv (ccn, npi, npi_source)
"""

from __future__ import annotations

import csv
import io
import re
import zipfile
from pathlib import Path

DOWNLOADS = Path("/data0/mrf-pricing-research/mrf/downloads.csv")
OUT = Path("/data0/mrf-pricing-research/crosswalk/ccn_to_npi.csv")

NPI_RE = re.compile(r"\b(\d{10})\b")
NPI_FIELD_NAMES = ("type_2_npi", "hospital_npi", "npi")


def npi_from_json(path: Path) -> list[str]:
    """Extract hospital NPI from JSON MRF metadata header.

    Streams only the leading bytes — full files can be 500+ MB.
    """
    with path.open("rb") as f:
        head = f.read(64 * 1024).decode("utf-8", "ignore")
    found = []
    # Look for "type_2_npi": [...] or "type_2_npi": "..."  (also fallbacks)
    for field in NPI_FIELD_NAMES:
        for m in re.finditer(
                rf'"{field}"\s*:\s*("(\d{{10}})"|\[([^\]]+)\])',
                head, re.IGNORECASE):
            single = m.group(2)
            arr = m.group(3) or ""
            if single:
                found.append(single)
            else:
                found.extend(NPI_RE.findall(arr))
        if found:
            break
    return list(dict.fromkeys(found))


def npi_from_csv(path: Path) -> list[str]:
    """Look for NPI in the metadata header (rows 1–2) of a CSV MRF.

    CMS v3.0 layout: row 1 = column headers (hospital_name, …, type_2_npi, …),
    row 2 = corresponding values. Multi-location facilities pipe-separate
    NPIs in a single cell.
    """
    try:
        with path.open("rb") as f:
            head_bytes = f.read(16384)
        head = head_bytes.decode("utf-8", "ignore").lstrip("﻿")
    except Exception:
        return []

    # Robust 2-row read using csv.reader to handle quoted multi-line cells.
    try:
        reader = csv.reader(head.splitlines())
        rows = []
        for row in reader:
            rows.append(row)
            if len(rows) >= 5:
                break
    except Exception:
        return []
    if len(rows) < 2:
        return []

    header = [c.strip().strip('"').lower() for c in rows[0]]
    values = rows[1]
    for field in NPI_FIELD_NAMES:
        if field in header:
            idx = header.index(field)
            if idx < len(values):
                vals = NPI_RE.findall(values[idx])
                if vals:
                    return list(dict.fromkeys(vals))

    # Fallback: any 10-digit number near a known field label in the head bytes
    head_lower = head.lower()
    for field in NPI_FIELD_NAMES:
        for m in re.finditer(rf"{field}[\"\']?\s*[,:=]?\s*[\"\']?(\d{{10}})",
                             head_lower):
            return [m.group(1)]
    return []


def sniff_and_extract(path: Path) -> tuple[list[str], str]:
    """Detect file format from magic bytes, route to format-specific extractor.

    Returns (npis, schema_version). Schema version comes from the MRF
    metadata's `version` field (e.g., "3.0.0", "2.0.0") when present;
    "?" when the format is unrecognized.

    Format detection uses content magic, not extension, since some
    hosts mis-name files (e.g., Kaiser .csv-named .zip files,
    UCSF .json-served .php).
    """
    with path.open("rb") as f:
        magic = f.read(8)

    # XLSX is a ZIP with a specific internal structure.
    is_zip = magic.startswith(b"PK\x03\x04")
    if is_zip:
        # Distinguish XLSX from generic data ZIP by looking for [Content_Types].xml
        try:
            with zipfile.ZipFile(path) as z:
                names = z.namelist()
                if "[Content_Types].xml" in names and any(
                        n.startswith("xl/") for n in names):
                    return _extract_from_xlsx(path)
                # Plain data ZIP — extract first CSV or JSON member
                members = [m for m in names
                           if m.lower().endswith((".csv", ".json"))]
                if not members:
                    return [], "?"
                with z.open(members[0]) as zf:
                    head = zf.read(64 * 1024)
        except Exception:
            return [], "?"
    else:
        with path.open("rb") as f:
            head = f.read(64 * 1024)

    text = head.decode("utf-8", "ignore").lstrip("﻿﻿").lstrip()
    if text.startswith("<"):
        return [], "html"  # HTML landing page; should already be filtered
    if text.startswith(("{", "[")):
        return _extract_from_json_text(text), _detect_json_version(text)
    npis = _extract_from_csv_text(text)
    return npis, _detect_csv_version(text)


def _extract_from_xlsx(path: Path) -> tuple[list[str], str]:
    """Use openpyxl in read-only mode to grab row 1 + row 2 of sheet 1.

    NPI may live in the metadata header just like CSV. Falls back
    silently if openpyxl isn't installed or the file is malformed.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return [], "xlsx"
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(c) if c is not None else "" for c in row])
            if len(rows) >= 5:
                break
        wb.close()
    except Exception:
        return [], "xlsx"
    if len(rows) < 2:
        return [], "xlsx"
    header = [c.strip().lower() for c in rows[0]]
    values = rows[1]
    version = "xlsx"
    if "version" in header:
        idx = header.index("version")
        if idx < len(values):
            version = values[idx] or version
    for field in NPI_FIELD_NAMES:
        if field in header:
            idx = header.index(field)
            if idx < len(values):
                vals = NPI_RE.findall(values[idx])
                if vals:
                    return list(dict.fromkeys(vals)), version
    return [], version


def _detect_json_version(text: str) -> str:
    m = re.search(r'"version"\s*:\s*"([^"]+)"', text, re.IGNORECASE)
    return m.group(1) if m else "?"


def _detect_csv_version(text: str) -> str:
    try:
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)[:2]
    except Exception:
        return "?"
    if len(rows) < 2:
        return "?"
    header = [c.strip().lower() for c in rows[0]]
    values = rows[1]
    if "version" in header:
        idx = header.index("version")
        if idx < len(values):
            return values[idx] or "?"
    return "?"


def _extract_from_json_text(text: str) -> list[str]:
    found = []
    for field in NPI_FIELD_NAMES:
        for m in re.finditer(
                rf'"{field}"\s*:\s*("(\d{{10}})"|\[([^\]]+)\])',
                text, re.IGNORECASE):
            single = m.group(2)
            arr = m.group(3) or ""
            if single:
                found.append(single)
            else:
                found.extend(NPI_RE.findall(arr))
        if found:
            break
    return list(dict.fromkeys(found))


def _extract_from_csv_text(text: str) -> list[str]:
    try:
        reader = csv.reader(io.StringIO(text))
        rows = []
        for row in reader:
            rows.append(row)
            if len(rows) >= 5:
                break
    except Exception:
        return []
    if len(rows) < 2:
        return []
    header = [c.strip().strip('"').lower() for c in rows[0]]
    values = rows[1]
    for field in NPI_FIELD_NAMES:
        if field in header:
            idx = header.index(field)
            if idx < len(values):
                vals = NPI_RE.findall(values[idx])
                if vals:
                    return list(dict.fromkeys(vals))
    text_lower = text.lower()
    for field in NPI_FIELD_NAMES:
        for m in re.finditer(rf"{field}[\"\']?\s*[,:=]?\s*[\"\']?(\d{{10}})",
                             text_lower):
            return [m.group(1)]
    return []


def main() -> None:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    pairs: dict[str, tuple[str, str, str]] = {}  # ccn -> (npi_or_blank, format, version)
    n_total = 0
    with DOWNLOADS.open() as f:
        for row in csv.DictReader(f):
            if row.get("status") != "ok":
                continue
            ccn = row["ccn"]
            path = Path(row["local_path"])
            if not path.exists() or ccn in pairs:
                continue
            n_total += 1
            try:
                npis, version = sniff_and_extract(path)
            except Exception as e:
                print(f"  [err] {ccn}: {e}")
                continue
            ext = path.suffix.lstrip(".").lower() or "?"
            if npis:
                pairs[ccn] = (npis[0], ext, version)
            else:
                # Record the negative — useful for downstream "why no NPI"
                pairs[ccn] = ("", ext, version)

    with OUT.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["ccn", "npi", "npi_source_format", "schema_version"])
        for ccn in sorted(pairs):
            npi, fmt, ver = pairs[ccn]
            w.writerow([ccn, npi, fmt, ver])

    n_with = sum(1 for p in pairs.values() if p[0])
    print(f"[out] {OUT}: {n_with}/{n_total} CCNs with NPI "
          f"({100*n_with/max(n_total,1):.1f}%)")

    # Coverage by schema version
    from collections import Counter
    ver_total = Counter(p[2] for p in pairs.values())
    ver_with_npi = Counter(p[2] for p in pairs.values() if p[0])
    print("\n[breakdown] NPI extraction by schema version:")
    for ver in sorted(ver_total, key=lambda v: -ver_total[v]):
        with_n = ver_with_npi.get(ver, 0)
        total = ver_total[ver]
        print(f"  v{ver:<10} {with_n:>4}/{total:<4} = "
              f"{100*with_n/total:.1f}%")


if __name__ == "__main__":
    main()
