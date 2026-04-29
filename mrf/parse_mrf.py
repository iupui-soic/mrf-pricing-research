#!/usr/bin/env python3
"""
parse_mrf.py
============
Schema-aware parser for the federal HPT MRF corpus. Reads every file
flagged status=ok in /data0/mrf-pricing-research/mrf/downloads.csv, detects format + schema
version, and emits long-format parquets.

**Resilience design**:
- Each hospital writes to its own `parts/<kind>_<state>_<ccn>.parquet`
  via an atomic `.tmp` → final rename. If the parser is killed mid-run,
  every completed hospital's data is permanently durable.
- Within a single file, rows are streamed and flushed to the writer in
  batches of `BATCH_ROWS` so even multi-GB inputs don't accumulate
  unbounded memory.
- A `done/<state>_<ccn>.flag` marker file enables resume: re-running
  skips hospitals that already finished.
- The append-mode `mrf_log.csv` gets one row per hospital as soon as
  the hospital finishes, with `flush()` after every write.
- A separate `concat_parts.py` merges the per-hospital parquets into
  unified `mrf_gross.parquet` / `mrf_negotiated.parquet` after the run.

**Schema adapters** (auto-detected by content sniffing, not extension):
  - v3 JSON          (`"version": "3.x"`, top-level
                       `standard_charge_information` array — streamed via ijson)
  - v3 / v2 CSV tall (3-row metadata header; data has explicit
                       `payer_name` and `plan_name` columns)
  - v3 / v2 CSV wide (3-row metadata header; per-payer column names
                       like `standard_charge|<PAYER>|<PLAN>|negotiated_dollar`)
  - XLSX             (uses openpyxl; same shape as CSV-tall/wide)
  - Kaiser legacy    (custom 2-row header; skipped — not §180 v2/v3)
  - ZIP              (extracts first inner csv/json on the fly)

Usage:
    .venv/bin/python -u mrf/parse_mrf.py            # full run, resume-aware
    .venv/bin/python -u mrf/parse_mrf.py --restart  # ignore done/ markers
"""

from __future__ import annotations

import argparse
import csv
import io
import os
import re
import sys
import time
import traceback
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterator

import ijson
import pyarrow as pa
import pyarrow.parquet as pq

OUT_DIR = Path("/data0/mrf-pricing-research/mrf/parsed")
PARTS_DIR = OUT_DIR / "parts"
DONE_DIR = OUT_DIR / "done"
LOG_CSV = OUT_DIR / "mrf_log.csv"
ERRORS_CSV = OUT_DIR / "mrf_errors.csv"
PROGRESS_LOG = OUT_DIR / "parse.log"

DOWNLOADS = Path("/data0/mrf-pricing-research/mrf/downloads.csv")

# Per-batch flush size (rows). Small enough to bound peak RAM, big
# enough to keep parquet row groups efficient.
BATCH_ROWS = 25_000

# ---------- Output schemas ----------

GROSS_SCHEMA = pa.schema([
    ("ccn", pa.string()),
    ("description", pa.string()),
    ("code", pa.string()),
    ("code_type", pa.string()),
    ("code2", pa.string()),
    ("code2_type", pa.string()),
    ("billing_class", pa.string()),
    ("setting", pa.string()),
    ("drug_unit", pa.string()),
    ("drug_type", pa.string()),
    ("modifiers", pa.string()),
    ("gross_charge", pa.float64()),
    ("discounted_cash", pa.float64()),
    ("standard_charge_min", pa.float64()),
    ("standard_charge_max", pa.float64()),
    ("schema_version", pa.string()),
    ("file_format", pa.string()),
    ("source_file", pa.string()),
])

NEG_SCHEMA = pa.schema([
    ("ccn", pa.string()),
    ("description", pa.string()),
    ("code", pa.string()),
    ("code_type", pa.string()),
    ("billing_class", pa.string()),
    ("setting", pa.string()),
    ("payer_name", pa.string()),
    ("plan_name", pa.string()),
    ("negotiated_dollar", pa.float64()),
    ("negotiated_percentage", pa.float64()),
    ("negotiated_algorithm", pa.string()),
    ("estimated_amount", pa.float64()),
    ("methodology", pa.string()),
    ("modifiers", pa.string()),
    ("schema_version", pa.string()),
    ("file_format", pa.string()),
    ("source_file", pa.string()),
])

LOG_FIELDS = [
    "ccn", "state", "source_file", "file_format", "schema_version",
    "n_gross", "n_negotiated", "duration_sec", "status", "error",
]

# ---------- Helpers ----------


def _to_float(s) -> float | None:
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return float(s)
    s = str(s).strip().replace(",", "").replace("$", "")
    if not s or s.lower() in ("na", "n/a", "none", "null", "-",
                              "not applicable"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _norm(s) -> str:
    if s is None:
        return ""
    return str(s).strip()


@dataclass
class FileLog:
    ccn: str
    state: str
    source_file: str
    file_format: str = ""
    schema_version: str = ""
    n_gross: int = 0
    n_negotiated: int = 0
    duration_sec: float = 0.0
    status: str = "pending"
    error: str = ""


@dataclass
class ParseContext:
    """Per-hospital parsing state. Schema adapters call emit_gross /
    emit_neg with row dicts; the context flushes to per-hospital
    parquet writers in BATCH_ROWS-sized chunks."""
    ccn: str
    schema_version: str
    file_format: str
    source_file: str
    log: FileLog
    gross_writer: pq.ParquetWriter
    neg_writer: pq.ParquetWriter
    _gross_buf: list = field(default_factory=list)
    _neg_buf: list = field(default_factory=list)

    def emit_gross(self, row: dict) -> None:
        self._gross_buf.append(row)
        self.log.n_gross += 1
        if len(self._gross_buf) >= BATCH_ROWS:
            self._flush_gross()

    def emit_neg(self, row: dict) -> None:
        self._neg_buf.append(row)
        self.log.n_negotiated += 1
        if len(self._neg_buf) >= BATCH_ROWS:
            self._flush_neg()

    def _flush_gross(self) -> None:
        if self._gross_buf:
            self.gross_writer.write_table(
                pa.Table.from_pylist(self._gross_buf, schema=GROSS_SCHEMA))
            self._gross_buf = []

    def _flush_neg(self) -> None:
        if self._neg_buf:
            self.neg_writer.write_table(
                pa.Table.from_pylist(self._neg_buf, schema=NEG_SCHEMA))
            self._neg_buf = []

    def flush_all(self) -> None:
        self._flush_gross()
        self._flush_neg()


# ---------- Format / schema detection ----------


def detect(path: Path) -> tuple[str, str, object]:
    """Return (file_format, schema_version, opened_resource)."""
    with path.open("rb") as f:
        magic = f.read(8)

    if magic.startswith(b"PK\x03\x04"):
        try:
            with zipfile.ZipFile(path) as z:
                if "[Content_Types].xml" in z.namelist():
                    return ("xlsx", _detect_xlsx_version(path), path)
                inner = next((n for n in z.namelist()
                              if n.lower().endswith((".csv", ".json"))), None)
                if not inner:
                    return ("unknown_zip", "?", None)
                with z.open(inner) as zf:
                    head = zf.read(64 * 1024)
        except Exception:
            return ("bad_zip", "?", None)
        return _detect_from_head(head, path, inner_name=inner)

    with path.open("rb") as f:
        head = f.read(64 * 1024)
    return _detect_from_head(head, path)


def _detect_from_head(head: bytes, path: Path,
                      inner_name: str | None = None
                      ) -> tuple[str, str, object]:
    text = head.decode("utf-8", "ignore").lstrip("﻿").lstrip()
    if text.startswith(("{", "[")):
        return ("json", _detect_json_version(text), (path, inner_name))
    if not text:
        return ("empty", "?", None)
    if text.lstrip().startswith("<"):
        return ("html", "?", None)

    try:
        reader = csv.reader(io.StringIO(text))
        rows = []
        for row in reader:
            rows.append(row)
            if len(rows) >= 3:
                break
    except Exception:
        return ("bad_csv", "?", None)

    if len(rows) < 2:
        return ("short_csv", "?", None)

    row0 = [c.strip().strip('"') for c in rows[0]]
    if "hospital_name" not in [c.lower() for c in row0]:
        return ("kaiser_legacy", "?", (path, inner_name))

    header_low = [c.strip().strip('"').lower() for c in rows[0]]
    values = rows[1] if len(rows) >= 2 else []
    version = "?"
    if "version" in header_low and len(values) > header_low.index("version"):
        version = (values[header_low.index("version")] or "?").strip()

    data_header = rows[2] if len(rows) >= 3 else []
    data_lower = [c.strip().strip('"').lower() for c in data_header]
    is_tall = "payer_name" in data_lower
    has_payer_cols = any("standard_charge|" in c and c.count("|") >= 3
                         for c in data_header)
    if is_tall:
        return ("csv_tall", version, (path, inner_name))
    if has_payer_cols:
        return ("csv_wide", version, (path, inner_name))
    return ("csv_minimal", version, (path, inner_name))


_JSON_VERSION_RE = re.compile(r'"version"\s*:\s*"([^"]+)"', re.IGNORECASE)


def _detect_json_version(text: str) -> str:
    m = _JSON_VERSION_RE.search(text)
    return m.group(1) if m else "?"


def _detect_xlsx_version(path: Path) -> str:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([("" if c is None else str(c)) for c in row])
            if len(rows) >= 2:
                break
        wb.close()
        if len(rows) < 2:
            return "?"
        h = [c.lower() for c in rows[0]]
        if "version" in h:
            i = h.index("version")
            if i < len(rows[1]):
                return rows[1][i] or "?"
    except Exception:
        pass
    return "?"


# ---------- CSV parsers ----------


def _open_csv_text(path: Path, inner_name: str | None
                   ) -> Iterator[list[str]]:
    if inner_name:
        with zipfile.ZipFile(path) as z:
            with z.open(inner_name) as zf:
                tio = io.TextIOWrapper(zf, encoding="utf-8", errors="replace",
                                       newline="")
                reader = csv.reader(tio)
                for row in reader:
                    yield row
    else:
        with path.open(encoding="utf-8", errors="replace", newline="") as f:
            first = f.read(1)
            if first != "﻿":
                f.seek(0)
            reader = csv.reader(f)
            for row in reader:
                yield row


_PAYER_COL_RE = re.compile(
    r"^standard_charge\|(?P<payer>[^|]+)\|(?P<plan>[^|]+)\|"
    r"(?P<field>negotiated_dollar|negotiated_percentage|"
    r"negotiated_algorithm|estimated_amount|methodology)$",
    re.IGNORECASE,
)


def _index_codes(header: list[str]) -> list[tuple[str, str]]:
    pairs = []
    for c in header:
        if not c:
            continue
        m = re.match(r"^code\|(\d+)$", c, re.IGNORECASE)
        if m:
            n = m.group(1)
            type_col = f"code|{n}|type"
            if type_col in header:
                pairs.append((c, type_col))
    return pairs


def parse_csv(ctx: ParseContext, path: Path, inner_name: str | None,
              wide: bool) -> None:
    rows_iter = _open_csv_text(path, inner_name)
    skipped = 0
    header_row = None
    for r in rows_iter:
        skipped += 1
        if skipped == 3:
            header_row = r
            break
    if header_row is None:
        ctx.log.status = "csv_too_short"
        return

    H = [c.strip().strip('"') for c in header_row]
    H_low = [c.lower() for c in H]
    code_pairs = _index_codes(H_low)

    def col(name):
        try:
            return H_low.index(name.lower())
        except ValueError:
            return None

    desc_i = col("description")
    bclass_i = col("billing_class")
    setting_i = col("setting")
    drug_u_i = col("drug_unit_of_measurement")
    drug_t_i = col("drug_type_of_measurement")
    mods_i = col("modifiers")
    gross_i = col("standard_charge|gross")
    cash_i = col("standard_charge|discounted_cash")
    min_i = col("standard_charge|min")
    max_i = col("standard_charge|max")
    payer_name_i = col("payer_name")
    plan_name_i = col("plan_name")
    neg_dollar_i = col("standard_charge|negotiated_dollar")
    neg_pct_i = col("standard_charge|negotiated_percentage")
    neg_algo_i = col("standard_charge|negotiated_algorithm")
    est_amt_i = col("estimated_amount")
    method_i = col("standard_charge|methodology") or col("methodology")

    wide_groups: dict = {}
    if wide:
        for i, c in enumerate(H):
            m = _PAYER_COL_RE.match(c)
            if m:
                key = (m.group("payer"), m.group("plan"))
                wide_groups.setdefault(key, {})[m.group("field").lower()] = i

    def _g(row, idx):
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    for row in rows_iter:
        if not row or all(not (c or "").strip() for c in row):
            continue
        try:
            description = _g(row, desc_i) or ""
            primary_code = primary_type = ""
            secondary_code = secondary_type = ""
            for code_col, type_col in code_pairs:
                code = _norm(_g(row, H_low.index(code_col)))
                ctype = _norm(_g(row, H_low.index(type_col)))
                if not code:
                    continue
                if not primary_code:
                    primary_code, primary_type = code, ctype
                elif not secondary_code:
                    secondary_code, secondary_type = code, ctype
                    break
            billing_class = _norm(_g(row, bclass_i))
            setting = _norm(_g(row, setting_i))
            drug_unit = _norm(_g(row, drug_u_i))
            drug_type = _norm(_g(row, drug_t_i))
            modifiers = _norm(_g(row, mods_i))
            gross = _to_float(_g(row, gross_i))
            cash = _to_float(_g(row, cash_i))
            sc_min = _to_float(_g(row, min_i))
            sc_max = _to_float(_g(row, max_i))

            if any(v is not None for v in (gross, cash, sc_min, sc_max)):
                ctx.emit_gross({
                    "ccn": ctx.ccn, "description": description,
                    "code": primary_code, "code_type": primary_type,
                    "code2": secondary_code, "code2_type": secondary_type,
                    "billing_class": billing_class, "setting": setting,
                    "drug_unit": drug_unit, "drug_type": drug_type,
                    "modifiers": modifiers,
                    "gross_charge": gross, "discounted_cash": cash,
                    "standard_charge_min": sc_min,
                    "standard_charge_max": sc_max,
                    "schema_version": ctx.schema_version,
                    "file_format": ctx.file_format,
                    "source_file": ctx.source_file,
                })

            if not wide:
                payer = _norm(_g(row, payer_name_i))
                plan = _norm(_g(row, plan_name_i))
                if payer or plan:
                    nd = _to_float(_g(row, neg_dollar_i))
                    np_ = _to_float(_g(row, neg_pct_i))
                    na = _norm(_g(row, neg_algo_i))
                    ea = _to_float(_g(row, est_amt_i))
                    md = _norm(_g(row, method_i))
                    if nd is not None or np_ is not None or na or ea is not None:
                        ctx.emit_neg({
                            "ccn": ctx.ccn, "description": description,
                            "code": primary_code, "code_type": primary_type,
                            "billing_class": billing_class, "setting": setting,
                            "payer_name": payer, "plan_name": plan,
                            "negotiated_dollar": nd,
                            "negotiated_percentage": np_,
                            "negotiated_algorithm": na,
                            "estimated_amount": ea,
                            "methodology": md,
                            "modifiers": modifiers,
                            "schema_version": ctx.schema_version,
                            "file_format": ctx.file_format,
                            "source_file": ctx.source_file,
                        })
            else:
                for (payer, plan), fields in wide_groups.items():
                    nd = _to_float(_g(row, fields.get("negotiated_dollar")))
                    np_ = _to_float(_g(row, fields.get("negotiated_percentage")))
                    na = _norm(_g(row, fields.get("negotiated_algorithm")))
                    ea = _to_float(_g(row, fields.get("estimated_amount")))
                    md = _norm(_g(row, fields.get("methodology")))
                    if nd is None and np_ is None and not na and ea is None:
                        continue
                    ctx.emit_neg({
                        "ccn": ctx.ccn, "description": description,
                        "code": primary_code, "code_type": primary_type,
                        "billing_class": billing_class, "setting": setting,
                        "payer_name": payer, "plan_name": plan,
                        "negotiated_dollar": nd,
                        "negotiated_percentage": np_,
                        "negotiated_algorithm": na,
                        "estimated_amount": ea,
                        "methodology": md,
                        "modifiers": modifiers,
                        "schema_version": ctx.schema_version,
                        "file_format": ctx.file_format,
                        "source_file": ctx.source_file,
                    })
        except Exception:
            continue


# ---------- JSON parser ----------


class _PrependedReader:
    """Wraps a binary file-like, prepending a small prefix before delegating."""

    def __init__(self, prefix: bytes, fp) -> None:
        self._prefix = prefix
        self._fp = fp

    def read(self, n: int = -1) -> bytes:
        if not self._prefix:
            return self._fp.read(n)
        if n is None or n < 0:
            data = self._prefix + self._fp.read()
            self._prefix = b""
            return data
        if n <= len(self._prefix):
            data, self._prefix = self._prefix[:n], self._prefix[n:]
            return data
        rem = self._prefix
        self._prefix = b""
        return rem + self._fp.read(n - len(rem))

    def close(self):
        return self._fp.close()


def parse_json(ctx: ParseContext, path: Path, inner_name: str | None) -> None:
    if inner_name:
        zf = zipfile.ZipFile(path)
        fp = zf.open(inner_name)
    else:
        zf = None
        fp = path.open("rb")
    # Strip optional UTF-8 BOM — ijson rejects it as "invalid char".
    sig = fp.read(3)
    if sig != b"\xef\xbb\xbf":
        fp = _PrependedReader(sig, fp)
    try:
        for item in ijson.items(fp, "standard_charge_information.item"):
            try:
                description = _norm(item.get("description"))
                billing_class = _norm(item.get("billing_class"))
                setting = _norm(item.get("setting"))
                drug_unit = _norm(item.get("drug_unit_of_measurement"))
                drug_type = _norm(item.get("drug_type_of_measurement"))
                modifiers = ",".join(str(m) for m in (item.get("modifiers") or []))

                codes = item.get("code_information") or []
                primary_code = primary_type = ""
                secondary_code = secondary_type = ""
                for ci in codes[:2]:
                    code = _norm(ci.get("code"))
                    ctype = _norm(ci.get("type"))
                    if not primary_code:
                        primary_code, primary_type = code, ctype
                    else:
                        secondary_code, secondary_type = code, ctype

                for sc in (item.get("standard_charges") or []):
                    gross = _to_float(sc.get("gross_charge"))
                    cash = _to_float(sc.get("discounted_cash"))
                    sc_min = _to_float(sc.get("minimum"))
                    sc_max = _to_float(sc.get("maximum"))
                    sc_setting = _norm(sc.get("setting") or setting)
                    sc_billing = _norm(sc.get("billing_class") or billing_class)

                    if any(v is not None for v in (gross, cash, sc_min, sc_max)):
                        ctx.emit_gross({
                            "ccn": ctx.ccn, "description": description,
                            "code": primary_code, "code_type": primary_type,
                            "code2": secondary_code, "code2_type": secondary_type,
                            "billing_class": sc_billing, "setting": sc_setting,
                            "drug_unit": drug_unit, "drug_type": drug_type,
                            "modifiers": modifiers,
                            "gross_charge": gross, "discounted_cash": cash,
                            "standard_charge_min": sc_min,
                            "standard_charge_max": sc_max,
                            "schema_version": ctx.schema_version,
                            "file_format": ctx.file_format,
                            "source_file": ctx.source_file,
                        })

                    for pi in (sc.get("payers_information") or []):
                        nd = _to_float(pi.get("standard_charge_dollar"))
                        np_ = _to_float(pi.get("standard_charge_percentage"))
                        na = _norm(pi.get("standard_charge_algorithm"))
                        ea = _to_float(pi.get("estimated_amount"))
                        md = _norm(pi.get("methodology"))
                        ctx.emit_neg({
                            "ccn": ctx.ccn, "description": description,
                            "code": primary_code, "code_type": primary_type,
                            "billing_class": sc_billing, "setting": sc_setting,
                            "payer_name": _norm(pi.get("payer_name")),
                            "plan_name": _norm(pi.get("plan_name")),
                            "negotiated_dollar": nd,
                            "negotiated_percentage": np_,
                            "negotiated_algorithm": na,
                            "estimated_amount": ea,
                            "methodology": md,
                            "modifiers": modifiers,
                            "schema_version": ctx.schema_version,
                            "file_format": ctx.file_format,
                            "source_file": ctx.source_file,
                        })
            except Exception:
                continue
    finally:
        try:
            fp.close()
        except Exception:
            pass
        if zf:
            zf.close()


# ---------- XLSX parser ----------


def parse_xlsx(ctx: ParseContext, path: Path) -> None:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active

        def text_rows():
            for row in ws.iter_rows(values_only=True):
                yield ["" if c is None else str(c) for c in row]

        gen = text_rows()
        for _ in range(2):
            next(gen, None)
        header = next(gen, None)
        if not header:
            ctx.log.status = "xlsx_too_short"
            wb.close()
            return

        H = list(header)
        H_low = [c.lower() for c in H]
        wide = any("standard_charge|" in (c or "")
                   and (c or "").count("|") >= 3 for c in H)

        def col(name):
            try:
                return H_low.index(name.lower())
            except ValueError:
                return None

        code_pairs = _index_codes(H_low)
        desc_i = col("description")
        bclass_i = col("billing_class")
        setting_i = col("setting")
        gross_i = col("standard_charge|gross")
        cash_i = col("standard_charge|discounted_cash")
        min_i = col("standard_charge|min")
        max_i = col("standard_charge|max")
        payer_name_i = col("payer_name")
        plan_name_i = col("plan_name")
        neg_dollar_i = col("standard_charge|negotiated_dollar")
        neg_pct_i = col("standard_charge|negotiated_percentage")
        neg_algo_i = col("standard_charge|negotiated_algorithm")
        method_i = col("standard_charge|methodology") or col("methodology")

        wide_groups: dict = {}
        if wide:
            for i, c in enumerate(H):
                m = _PAYER_COL_RE.match(c or "")
                if m:
                    key = (m.group("payer"), m.group("plan"))
                    wide_groups.setdefault(key, {})[m.group("field").lower()] = i

        def _g(row, idx):
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        for row in gen:
            if not row or all(not (c or "").strip() for c in row):
                continue
            try:
                description = _g(row, desc_i) or ""
                primary_code = primary_type = ""
                for code_col, type_col in code_pairs:
                    code = _norm(_g(row, H_low.index(code_col)))
                    ctype = _norm(_g(row, H_low.index(type_col)))
                    if code:
                        primary_code, primary_type = code, ctype
                        break
                billing_class = _norm(_g(row, bclass_i))
                setting = _norm(_g(row, setting_i))
                gross = _to_float(_g(row, gross_i))
                cash = _to_float(_g(row, cash_i))
                sc_min = _to_float(_g(row, min_i))
                sc_max = _to_float(_g(row, max_i))

                if any(v is not None for v in (gross, cash, sc_min, sc_max)):
                    ctx.emit_gross({
                        "ccn": ctx.ccn, "description": description,
                        "code": primary_code, "code_type": primary_type,
                        "code2": "", "code2_type": "",
                        "billing_class": billing_class, "setting": setting,
                        "drug_unit": "", "drug_type": "", "modifiers": "",
                        "gross_charge": gross, "discounted_cash": cash,
                        "standard_charge_min": sc_min,
                        "standard_charge_max": sc_max,
                        "schema_version": ctx.schema_version,
                        "file_format": ctx.file_format,
                        "source_file": ctx.source_file,
                    })

                if not wide:
                    payer = _norm(_g(row, payer_name_i))
                    plan = _norm(_g(row, plan_name_i))
                    if payer or plan:
                        nd = _to_float(_g(row, neg_dollar_i))
                        np_ = _to_float(_g(row, neg_pct_i))
                        na = _norm(_g(row, neg_algo_i))
                        md = _norm(_g(row, method_i))
                        if nd is not None or np_ is not None or na:
                            ctx.emit_neg({
                                "ccn": ctx.ccn, "description": description,
                                "code": primary_code, "code_type": primary_type,
                                "billing_class": billing_class, "setting": setting,
                                "payer_name": payer, "plan_name": plan,
                                "negotiated_dollar": nd,
                                "negotiated_percentage": np_,
                                "negotiated_algorithm": na,
                                "estimated_amount": None,
                                "methodology": md,
                                "modifiers": "",
                                "schema_version": ctx.schema_version,
                                "file_format": ctx.file_format,
                                "source_file": ctx.source_file,
                            })
                else:
                    for (payer, plan), fields in wide_groups.items():
                        nd = _to_float(_g(row, fields.get("negotiated_dollar")))
                        np_ = _to_float(_g(row, fields.get("negotiated_percentage")))
                        na = _norm(_g(row, fields.get("negotiated_algorithm")))
                        md = _norm(_g(row, fields.get("methodology")))
                        if nd is None and np_ is None and not na:
                            continue
                        ctx.emit_neg({
                            "ccn": ctx.ccn, "description": description,
                            "code": primary_code, "code_type": primary_type,
                            "billing_class": billing_class, "setting": setting,
                            "payer_name": payer, "plan_name": plan,
                            "negotiated_dollar": nd,
                            "negotiated_percentage": np_,
                            "negotiated_algorithm": na,
                            "estimated_amount": None,
                            "methodology": md,
                            "modifiers": "",
                            "schema_version": ctx.schema_version,
                            "file_format": ctx.file_format,
                            "source_file": ctx.source_file,
                        })
            except Exception:
                continue
        wb.close()
    except Exception as e:
        ctx.log.status = "xlsx_failed"
        ctx.log.error = str(e)[:200]


# ---------- Driver ----------


def list_files() -> list[dict]:
    rows = []
    with DOWNLOADS.open() as f:
        for r in csv.DictReader(f):
            if r.get("status") == "ok" and Path(r["local_path"]).exists():
                rows.append(r)
    return rows


def part_paths(state: str, ccn: str) -> tuple[Path, Path, Path]:
    """Return (gross_part, neg_part, done_marker) for this hospital."""
    return (
        PARTS_DIR / f"gross_{state}_{ccn}.parquet",
        PARTS_DIR / f"neg_{state}_{ccn}.parquet",
        DONE_DIR / f"{state}_{ccn}.flag",
    )


def parse_one(d: dict, log_writer, log_fp, *, restart: bool) -> FileLog:
    ccn = d["ccn"]
    state = d["state"]
    path = Path(d["local_path"])
    log = FileLog(ccn=ccn, state=state, source_file=str(path))

    gross_part, neg_part, done_flag = part_paths(state, ccn)

    # Resume: skip if previously completed
    if not restart and done_flag.exists():
        log.status = "resume_skip"
        return log

    fmt, version, resource = detect(path)
    log.file_format = fmt
    log.schema_version = version

    if fmt in {"empty", "html", "unknown_zip", "bad_zip", "bad_csv",
               "short_csv", "kaiser_legacy"}:
        log.status = f"skip:{fmt}"
        log.duration_sec = 0.0
        # Touch a done marker so resume doesn't re-attempt
        done_flag.touch()
        return log

    # Open per-hospital writers as .tmp
    gtmp = gross_part.with_suffix(".parquet.tmp")
    ntmp = neg_part.with_suffix(".parquet.tmp")
    for p in (gtmp, ntmp):
        if p.exists():
            p.unlink()
    gw = pq.ParquetWriter(gtmp, GROSS_SCHEMA, compression="snappy")
    nw = pq.ParquetWriter(ntmp, NEG_SCHEMA, compression="snappy")
    ctx = ParseContext(ccn=ccn, schema_version=version, file_format=fmt,
                       source_file=str(path), log=log,
                       gross_writer=gw, neg_writer=nw)

    t0 = time.monotonic()
    try:
        if fmt == "json":
            p, inner = resource
            parse_json(ctx, p, inner)
        elif fmt in ("csv_tall", "csv_minimal"):
            p, inner = resource
            parse_csv(ctx, p, inner, wide=False)
        elif fmt == "csv_wide":
            p, inner = resource
            parse_csv(ctx, p, inner, wide=True)
        elif fmt == "xlsx":
            parse_xlsx(ctx, path)
        else:
            log.status = f"skip:{fmt}"
        ctx.flush_all()
        if not log.status.startswith("skip:") and not log.status.startswith("xlsx_failed"):
            log.status = "ok"
    except Exception as e:
        log.status = "parse_failed"
        log.error = (str(e) + "\n" + traceback.format_exc()[-400:])[:1000]
    finally:
        gw.close()
        nw.close()
        log.duration_sec = round(time.monotonic() - t0, 2)

    # Atomic finalization (only on ok / xlsx_failed-with-partial)
    if log.status == "ok":
        gtmp.replace(gross_part)
        ntmp.replace(neg_part)
        done_flag.touch()
    else:
        # Delete the .tmp parquets — they're partial / unwanted
        gtmp.unlink(missing_ok=True)
        ntmp.unlink(missing_ok=True)
        if log.status.startswith("skip:") or log.status.startswith("xlsx_failed"):
            done_flag.touch()  # don't retry on next run

    return log


def write_log_row(log_writer, log_fp, log: FileLog) -> None:
    log_writer.writerow({
        "ccn": log.ccn, "state": log.state,
        "source_file": log.source_file,
        "file_format": log.file_format,
        "schema_version": log.schema_version,
        "n_gross": log.n_gross,
        "n_negotiated": log.n_negotiated,
        "duration_sec": log.duration_sec,
        "status": log.status,
        "error": log.error[:300],
    })
    log_fp.flush()
    os.fsync(log_fp.fileno())


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--restart", action="store_true",
                    help="ignore done/ markers and re-parse everything")
    args = ap.parse_args()

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    PARTS_DIR.mkdir(parents=True, exist_ok=True)
    DONE_DIR.mkdir(parents=True, exist_ok=True)

    files = list_files()
    n_total = len(files)

    # Open log in append mode for resume; reset on --restart
    log_mode = "w" if args.restart or not LOG_CSV.exists() else "a"
    log_fp = LOG_CSV.open(log_mode, newline="")
    log_writer = csv.DictWriter(log_fp, fieldnames=LOG_FIELDS)
    if log_mode == "w":
        log_writer.writeheader()
        log_fp.flush()

    print(f"[plan] {n_total} files; restart={args.restart}", flush=True)
    print(f"[plan] parts dir: {PARTS_DIR}", flush=True)
    print(f"[plan] log: {LOG_CSV}", flush=True)

    cum_gross = cum_neg = 0
    n_ok = n_skip = n_fail = n_resume = 0
    t_start = time.monotonic()
    try:
        for i, d in enumerate(files, 1):
            log = parse_one(d, log_writer, log_fp, restart=args.restart)

            if log.status == "ok":
                n_ok += 1
                cum_gross += log.n_gross
                cum_neg += log.n_negotiated
                write_log_row(log_writer, log_fp, log)
            elif log.status == "resume_skip":
                n_resume += 1
                # Don't re-write log row; the prior row is still there
            elif log.status.startswith("skip:"):
                n_skip += 1
                write_log_row(log_writer, log_fp, log)
            else:
                n_fail += 1
                write_log_row(log_writer, log_fp, log)

            elapsed = time.monotonic() - t_start
            rate = i / max(elapsed, 0.001)
            eta_min = (n_total - i) / max(rate, 0.001) / 60
            print(f"[{i:3d}/{n_total}] {log.status:14} ccn={log.ccn} "
                  f"({log.state}) {log.file_format:<12} v{log.schema_version:<6} "
                  f"gross={log.n_gross:>7,} neg={log.n_negotiated:>9,} "
                  f"in {log.duration_sec:>6.1f}s | "
                  f"cum: ok={n_ok} skip={n_skip} fail={n_fail} resumed={n_resume} "
                  f"| ETA {eta_min:.1f} min",
                  flush=True)
    finally:
        log_fp.close()

    print(f"\n[done] {n_total} files: ok={n_ok}, skipped={n_skip}, "
          f"failed={n_fail}, resumed={n_resume}", flush=True)
    print(f"       cumulative rows: gross={cum_gross:,} neg={cum_neg:,}", flush=True)
    print(f"       parts at: {PARTS_DIR}", flush=True)
    print(f"       run `python mrf/concat_parts.py` to merge into "
          f"mrf_gross.parquet + mrf_negotiated.parquet", flush=True)
    return 0 if n_fail == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
